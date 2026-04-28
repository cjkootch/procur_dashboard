"""
Build the master Caribbean fuel supplier database.
v2: uses dr_fuel_awards_v2.json (per-award fuel filter, real values, real buyers)
"""
import json
import re
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path("/home/claude/caribbean_fuel")
OUTPUTS_DIR = Path("/mnt/user-data/outputs")
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

# ─── Load source data ──────────────────────────────────────────────────────────

dr = json.load(open(OUT / "dr_fuel_awards_v2.json"))
jm = json.load(open(OUT / "gojep_fuel_awards.json"))

# ─── Helpers ───────────────────────────────────────────────────────────────────

def normalize_supplier_name(name: str) -> str:
    if not name: return ""
    n = name.lower().strip().rstrip("\t")
    n = re.sub(r"\s+", " ", n)
    suffixes = [
        ", srl", " srl", ", s.r.l.", " s.r.l.", ", s.r.l", ", s.a.s.", " sas",
        ", sas", ", s.a.", " s.a.", " s. a.", ", s.a", " s.a", ", sa", " sa",
        ", inc", " inc.", " inc", ", llc", " llc", ", ltd.", " ltd.", " ltd",
        " ltda", " corp.", " corp", ", c. por a.", " c. por a.", " c por a",
        " s de rl", ", s de rl", ", se", " se",
    ]
    changed = True
    while changed:
        changed = False
        for sfx in suffixes:
            if n.endswith(sfx):
                n = n[:-len(sfx)].strip()
                changed = True
    n = re.sub(r"[,.]+$", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def parse_date(s):
    if not s: return None
    s = str(s)
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if m: return m.group(1)
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m: return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None


def fuel_category_from_unspsc(codes):
    cats = set()
    for c in codes or []:
        c = str(c)
        if c in {"15101502", "15101505"}: cats.add("diesel")
        elif c in {"15101503", "15101506"}: cats.add("gasoline")
        elif c in {"15101504", "15101510"}: cats.add("aviation")
        elif c == "15101508": cats.add("crude")
        elif c in {"15101509", "15101516"}: cats.add("marine_bunker")
        elif c == "15101512": cats.add("lpg")
        elif c == "15101517": cats.add("lng")
        elif c == "15101701": cats.add("heating_oil")
        elif c == "15101702": cats.add("heavy_fuel_oil")
    return sorted(cats)


def fuel_category_from_title(title):
    t = (title or "").lower()
    cats = set()
    if "diesel" in t or "gasoil" in t or "ultra low sulphur" in t: cats.add("diesel")
    if "gasoline" in t or "unleaded" in t or "petrol" in t: cats.add("gasoline")
    if "jet" in t or "aviation" in t: cats.add("aviation")
    if "lpg" in t or "liquid petroleum gas" in t or "liquefied petroleum" in t: cats.add("lpg")
    if "bunker fuel" in t or "marine fuel" in t: cats.add("marine_bunker")
    if "fuel oil" in t and "diesel" not in t: cats.add("heavy_fuel_oil")
    if "fuel for generator" in t and not cats: cats.add("diesel")
    if not cats and "fuel" in t: cats.add("unspecified")
    return sorted(cats)


# Currency conversion (approximate, current spot rates)
DOP_USD = 60.0
JMD_USD = 158.0

def to_usd(amount, currency):
    if amount is None: return None
    try: amt = float(amount)
    except (ValueError, TypeError): return None
    if currency in ("USD", "US$", "USD$"): return amt
    if currency == "DOP": return round(amt / DOP_USD, 2)
    if currency == "JMD": return round(amt / JMD_USD, 2)
    return None


# International majors / regional principals (manually curated)
INTL_MAJORS = {
    "totalenergies": "international_major",
    "total energies": "international_major",
    "total ": "international_major",
    "rubis": "international_major",
    "shell": "international_major",
    "exxonmobil": "international_major",
    "chevron": "international_major",
    "trafigura": "international_major",
    "vitol": "international_major",
    "glencore": "international_major",
    "puma energy": "international_major",
}
REGIONAL_MAJORS = {
    "sol jamaica": "regional_major",
    "sol caribbean": "regional_major",
    "isla dominicana de petroleo": "regional_major",
    "petromovil": "regional_major",
    "sigma petroleum": "regional_major",
    "sunix petroleum": "regional_major",
    "gulfstream petroleum": "regional_major",
    "next dominicana": "regional_major",
    "eco petroleo dominicana": "regional_major",
    "future energy source": "local_principal",
    "igl limited": "regional_major",  # IGL is large LPG distributor
    "rapid onsite refuelling": "local_distributor",
}

def classify_supplier(canonical_name: str) -> str:
    n = canonical_name.lower()
    for key, label in INTL_MAJORS.items():
        if key in n:
            return label
    for key, label in REGIONAL_MAJORS.items():
        if key in n:
            return label
    if any(k in n for k in [", srl", " srl", "diesel", "estación", "estacion", "gas station"]):
        return "local_distributor"
    if "petroleum" in n or "petrolera" in n or "petróleo" in n or "petroleo" in n:
        return "local_principal"
    return "local_other"


# ─── Flatten DR records ─────────────────────────────────────────────────────────

awards = []

for r in dr:
    cur = r.get("value_currency") or "DOP"
    val = r.get("value_native")
    awards.append({
        "country": "DOM",
        "source_portal": "DR_DGCP_OCDS",
        "ocid": r.get("ocid"),
        "tender_id": r.get("tender_id"),
        "tender_title": r.get("tender_title"),
        "buyer": r.get("buyer") or "UNKNOWN",
        "buyer_country": "DOM",
        "supplier_name": (r.get("supplier_name") or "").strip(),
        "supplier_name_normalized": normalize_supplier_name(r.get("supplier_name")),
        "supplier_id": r.get("supplier_id"),
        "award_id": r.get("award_id"),
        "award_date": parse_date(r.get("award_date")),
        "award_status": r.get("award_status"),
        "value_native": val,
        "value_currency": cur,
        "value_usd": to_usd(val, cur),
        "fuel_categories": fuel_category_from_unspsc(r.get("fuel_unspsc_codes")),
        "unspsc_codes": r.get("fuel_unspsc_codes") or [],
    })

# ─── Flatten Jamaica records ────────────────────────────────────────────────────

for r in jm:
    title = r.get("title", "") or ""
    # filter out the obvious non-fuel false positive
    tl = title.lower()
    if "compressor" in tl and "diesel" in tl:
        continue
    cur = r.get("currency") or "JMD"
    try:
        val = float(r.get("contract_price")) if r.get("contract_price") else None
    except (ValueError, TypeError):
        val = None
    awards.append({
        "country": "JAM",
        "source_portal": "GOJEP",
        "ocid": None,
        "tender_id": str(r.get("resource_id")),
        "tender_title": title,
        "buyer": r.get("buyer") or "UNKNOWN",
        "buyer_country": "JAM",
        "supplier_name": (r.get("awardee") or "").strip(),
        "supplier_name_normalized": normalize_supplier_name(r.get("awardee")),
        "supplier_id": None,
        "award_id": str(r.get("resource_id")),
        "award_date": parse_date(r.get("award_date")),
        "award_status": "active",
        "value_native": val,
        "value_currency": cur,
        "value_usd": to_usd(val, cur),
        "fuel_categories": fuel_category_from_title(title),
        "unspsc_codes": r.get("cpv_codes") or [],
    })

awards = [a for a in awards if a["supplier_name"]]
print(f"Total awards: {len(awards):,}")

# Save raw
(OUT / "caribbean_fuel_awards_master.json").write_text(
    json.dumps(awards, indent=2, ensure_ascii=False, default=str)
)


# ─── Build supplier roll-up ────────────────────────────────────────────────────

suppliers = defaultdict(lambda: {
    "supplier_name_canonical": "",
    "name_variants": set(),
    "countries_served": set(),
    "buyers": Counter(),
    "fuel_categories": set(),
    "awards_count": 0,
    "awards_value_usd_total": 0.0,
    "first_award_date": None,
    "last_award_date": None,
    "largest_award_usd": 0.0,
    "supplier_ids": set(),
    "source_portals": set(),
})

for a in awards:
    key = a["supplier_name_normalized"]
    if not key: continue
    s = suppliers[key]
    if not s["supplier_name_canonical"] or len(a["supplier_name"]) > len(s["supplier_name_canonical"]):
        s["supplier_name_canonical"] = a["supplier_name"]
    s["name_variants"].add(a["supplier_name"])
    s["countries_served"].add(a["country"])
    s["buyers"][a["buyer"]] += 1
    s["fuel_categories"].update(a["fuel_categories"])
    s["awards_count"] += 1
    if a["value_usd"]:
        s["awards_value_usd_total"] += a["value_usd"]
        s["largest_award_usd"] = max(s["largest_award_usd"], a["value_usd"])
    if a["award_date"]:
        if not s["first_award_date"] or a["award_date"] < s["first_award_date"]:
            s["first_award_date"] = a["award_date"]
        if not s["last_award_date"] or a["award_date"] > s["last_award_date"]:
            s["last_award_date"] = a["award_date"]
    if a["supplier_id"]: s["supplier_ids"].add(a["supplier_id"])
    s["source_portals"].add(a["source_portal"])

for key, s in suppliers.items():
    s["supplier_type"] = classify_supplier(s["supplier_name_canonical"])

print(f"Unique suppliers: {len(suppliers):,}")

# ─── Build dataframes ──────────────────────────────────────────────────────────

supplier_rows = []
for key, s in sorted(suppliers.items(), key=lambda x: -x[1]["awards_value_usd_total"]):
    supplier_rows.append({
        "rank_by_value": 0,
        "supplier_name": s["supplier_name_canonical"],
        "supplier_type": s["supplier_type"],
        "countries_served": ", ".join(sorted(s["countries_served"])),
        "fuel_categories": ", ".join(sorted(s["fuel_categories"])),
        "awards_count": s["awards_count"],
        "awards_value_usd_total": round(s["awards_value_usd_total"]) if s["awards_value_usd_total"] else None,
        "largest_award_usd": round(s["largest_award_usd"]) if s["largest_award_usd"] else None,
        "first_award_date": s["first_award_date"],
        "last_award_date": s["last_award_date"],
        "unique_buyers": len(s["buyers"]),
        "top_buyer": s["buyers"].most_common(1)[0][0] if s["buyers"] else None,
        "top_buyer_awards": s["buyers"].most_common(1)[0][1] if s["buyers"] else None,
        "name_variants_count": len(s["name_variants"]),
        "source_portals": ", ".join(sorted(s["source_portals"])),
    })
for i, r in enumerate(supplier_rows, 1):
    r["rank_by_value"] = i
df_suppliers = pd.DataFrame(supplier_rows)

df_awards = pd.DataFrame([
    {
        "country": a["country"],
        "source_portal": a["source_portal"],
        "buyer": a["buyer"],
        "supplier_name": a["supplier_name"],
        "tender_title": (a["tender_title"] or "")[:200],
        "fuel_categories": ", ".join(a["fuel_categories"]),
        "award_date": a["award_date"],
        "value_native": a["value_native"],
        "value_currency": a["value_currency"],
        "value_usd": a["value_usd"],
        "tender_id": a["tender_id"],
        "ocid": a["ocid"],
    }
    for a in awards
]).sort_values(["country", "value_usd"], ascending=[True, False], na_position="last")

# Heatmap (suppliers with 2+ awards)
fuel_cats = ["diesel", "gasoline", "lpg", "marine_bunker", "aviation",
             "heating_oil", "heavy_fuel_oil", "lng", "crude", "unspecified"]
heatmap_rows = []
for key, s in suppliers.items():
    if s["awards_count"] < 2: continue
    row = {
        "supplier_name": s["supplier_name_canonical"],
        "type": s["supplier_type"],
        "awards": s["awards_count"],
        "value_usd_total": round(s["awards_value_usd_total"]) if s["awards_value_usd_total"] else None,
        "DOM": "✓" if "DOM" in s["countries_served"] else "",
        "JAM": "✓" if "JAM" in s["countries_served"] else "",
    }
    for c in fuel_cats:
        row[c] = "✓" if c in s["fuel_categories"] else ""
    heatmap_rows.append(row)
df_heatmap = pd.DataFrame(heatmap_rows).sort_values("value_usd_total", ascending=False, na_position="last")

# Buyers
buyer_agg = defaultdict(lambda: {
    "country": "", "awards_count": 0, "value_usd_total": 0.0,
    "suppliers_used": set(), "fuel_categories": set(),
})
for a in awards:
    b = buyer_agg[a["buyer"]]
    b["country"] = a["country"]
    b["awards_count"] += 1
    if a["value_usd"]: b["value_usd_total"] += a["value_usd"]
    b["suppliers_used"].add(a["supplier_name_normalized"])
    b["fuel_categories"].update(a["fuel_categories"])
buyer_rows = []
for name, b in sorted(buyer_agg.items(), key=lambda x: -x[1]["value_usd_total"]):
    buyer_rows.append({
        "buyer": name or "UNKNOWN",
        "country": b["country"],
        "awards_count": b["awards_count"],
        "value_usd_total": round(b["value_usd_total"]) if b["value_usd_total"] else None,
        "unique_suppliers": len(b["suppliers_used"]),
        "fuel_categories": ", ".join(sorted(b["fuel_categories"])),
    })
df_buyers = pd.DataFrame(buyer_rows)

# ─── Write Excel ────────────────────────────────────────────────────────────────

out_xlsx = OUTPUTS_DIR / "caribbean_fuel_supplier_database.xlsx"

cover = pd.DataFrame([
    ["Caribbean Fuel Supplier Database", ""],
    ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")],
    ["Coverage", "Dominican Republic + Jamaica fuel supply contracts, 2021–2026"],
    ["Sources", "DR DGCP (OCDS bulk download) | Jamaica GOJEP (HTML+PDF scrape)"],
    ["Filter", "UNSPSC class 1510 fuel-supply codes (DR) + CPV 09xxx + title heuristics (JAM)"],
    ["", ""],
    ["── Headline numbers ──", ""],
    ["Total awards", f"{len(awards):,}"],
    ["Total unique suppliers", f"{len(suppliers):,}"],
    ["Suppliers with 2+ awards", f"{sum(1 for s in suppliers.values() if s['awards_count'] >= 2):,}"],
    ["Total contract value (USD)", f"${sum(a.get('value_usd') or 0 for a in awards):,.0f}"],
    ["", ""],
    ["── How to use this ──", ""],
    ["1. Suppliers tab", "Sorted by total contract value. Filter by supplier_type to find international majors, regional majors, or local distributors."],
    ["2. Heatmap tab", "Capability matrix — which suppliers can do which fuel categories in which countries. Use this when scoping a specific tender response."],
    ["3. Buyers tab", "Caribbean public fuel buyers ranked by spend. These are VTC's potential customers."],
    ["4. Awards (raw)", "Every individual contract award. Use for forensic deep-dives into specific suppliers or buyers."],
    ["", ""],
    ["── Caveats ──", ""],
    ["DR coverage", "Comprehensive (~6,000 fuel contracts via OCDS bulk download)"],
    ["Jamaica coverage", "Limited to keyword matches on GOJEP search ('fuel', 'diesel', 'petroleum', etc.); awards <5 years old"],
    ["Currency conversion", "Approximate spot rates: DOP/USD=60, JMD/USD=158. Refresh for active deal valuation."],
    ["Supplier type classification", "Heuristic — verify before outreach. International majors and named regional players are accurate; 'local_distributor' / 'local_other' may need manual review."],
    ["Missing markets", "Trinidad, Barbados, Bahamas, Guyana, Suriname not yet scraped (no OCDS feeds available; portal scrapers required)"],
], columns=["Field", "Value"])

with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
    cover.to_excel(w, sheet_name="Cover", index=False)
    df_suppliers.to_excel(w, sheet_name="Suppliers", index=False)
    df_heatmap.to_excel(w, sheet_name="Heatmap", index=False)
    df_buyers.to_excel(w, sheet_name="Buyers", index=False)
    df_awards.head(50000).to_excel(w, sheet_name="Awards (raw)", index=False)

# Format
wb = load_workbook(out_xlsx)
for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    # Auto-width
    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
wb.save(out_xlsx)

print(f"\nWrote {out_xlsx}")
print(f"Size: {out_xlsx.stat().st_size:,} bytes")

# Console summary
print("\n=== TOP 25 SUPPLIERS BY TOTAL USD AWARD VALUE ===")
for r in supplier_rows[:25]:
    val = r['awards_value_usd_total'] or 0
    print(f"  ${val:>13,} | {r['awards_count']:4d}× | {r['supplier_type']:22s} | {r['countries_served']:8s} | {r['supplier_name'][:50]}")

print("\n=== TOP 20 BUYERS BY USD VALUE ===")
for r in buyer_rows[:20]:
    val = r['value_usd_total'] or 0
    name = (r['buyer'] or 'UNKNOWN')[:55]
    print(f"  ${val:>13,} | {r['country']} | {name:55s} | {r['awards_count']:4d} awards | {r['unique_suppliers']:3d} suppliers")
